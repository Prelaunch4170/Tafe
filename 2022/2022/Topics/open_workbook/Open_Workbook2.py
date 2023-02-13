import json
import os

from openpyxl import load_workbook

wb = load_workbook('sample.xlsx')

sheet = wb.active
products = {}

for row in sheet.iter_rows(min_row=2, max_row=6, min_col=4, max_col=7, values_only=True):
    product_id = row[0]
    product = {
        "parent": row[1],
        "title": row[2],
        "category": row[3],
    }

    # create the dictionary - key - product id and value
    products[product_id] = product

print(json.dumps(products))

os.system(r"sample.xlsx")
